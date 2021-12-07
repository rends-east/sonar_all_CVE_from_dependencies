from decimal import Decimal
import json
import requests
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
Decimal('0.1')


def adjust_column_width_from_col(ws, min_row, min_col, max_col):
    column_widths = []
    for i, col in \
            enumerate(
                ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)
            ):
        for cell in col:
            value = cell.value
            if value is not None:
                if isinstance(value, str) is False:
                    value = str(value)
                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))
    for i, width in enumerate(column_widths):
        col_name = get_column_letter(min_col + i)
        value = column_widths[i] + 2
        ws.column_dimensions[col_name].width = value

def find_all(string, substr):
    start = 0
    while True:
        start = string.find(substr, start)
        if start == -1: return
        yield start
        start += len(substr)

def xl_out(Project_name, workbook, CVSS_Best_Scores, Count_of_vulns):
    try:
        wb = load_workbook(filename = workbook)
        ogl = wb.get_sheet_by_name("Content")
    except:
        wb = Workbook()
        wb.remove(wb['Sheet'])
        ogl = wb.create_sheet("Content")


    if(len(Project_name) > 31):
        i = ogl.max_row
        ogl['A' + str(i)] = Project_name
        ogl['B' + str(i)] = Project_name.split('.')[-1]
        Project_name = Project_name.split('.')[-1]

    if len(CVSS_Best_Scores) == 0:
        return
    try:
        ws1 = wb.get_sheet_by_name(Project_name)
    except:
        ws1 = wb.create_sheet(Project_name)
    ws1['A1'] = "Filename"
    ws1['B1'] = "CVSS max score"
    ws1['D1'] = "Summary of CVSS"
    ws1['C1'] = 'Number of CVE'
    k = 2
    for i in sorted(CVSS_Best_Scores.items(), key=lambda CVSS: (CVSS[1]),  reverse = True):
        ws1['A' + str(k)]= i[0]
        ws1['B' + str(k)] = i[1]
        ws1['C' + str(k)] = Count_of_vulns[i[0]][0]
        ws1['D' + str(k)] = Count_of_vulns[i[0]][1]
        k += 1
    adjust_column_width_from_col(ws1, 1, 1, 4)
    wb.save(filename = workbook)
def get_all_project_names(sonar_auth, url):
    all_project = requests.get(url + "/api/projects/search?qualifiers=TRK&ps=500", cookies=sonar_auth.cookies)
    json1 = json.loads(all_project.text)
    for project in json1['components']:
        yield [project['key'], project['name']]

def get_all_files_and_CVSS_from_project(project_name, sonar_auth, CVSS_Scores, url):
    file = ''
    CVSS = ''
    all_files = requests.get(url+ "/api/issues/search?componentKeys="+ project_name +"&owaspTop10=a9&resolved=false&types=VULNERABILITY", cookies=sonar_auth.cookies)
    json_issues = json.loads(all_files.text)
    for issue in json_issues['issues']:
        pipes = list(find_all(issue['message'], '|'))
        if len(pipes) > 0:
            file = issue['message'][10:pipes[0]]
            CVSS = issue['message'][pipes[1] + 14:pipes[2]]
            CVSS_Scores.append([file, CVSS])


def get_best_CVSS(CVSS_Scores, Count_of_vulns, CVSS_Best_Scores):
    for i in CVSS_Scores:
        try:
            Count_of_vulns[i[0]][0] += 1
            Count_of_vulns[i[0]][1] += Decimal(i[1])
        except:
            Count_of_vulns[i[0]] = []
            Count_of_vulns[i[0]].append(1)
            # print(i[1], i[0])
            try:
                Count_of_vulns[i[0]].append(Decimal(i[1]))
            except:
                print("Invalid CVSS detected!")
    for i in CVSS_Scores:
        try:
            CVSS_Best_Scores[i[0]] = max(float(i[1]), CVSS_Best_Scores[i[0]])
        except:
            CVSS_Best_Scores[i[0]] = 0.0
            try:
                CVSS_Best_Scores[i[0]] = max(float(i[1]), CVSS_Best_Scores[i[0]])
            except:
                pass
